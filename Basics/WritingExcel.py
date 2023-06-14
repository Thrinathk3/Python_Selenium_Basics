import openpyxl

# File-->WorkBook-->Sheets-->Rows-->Cells

# file = "D:\\programs work space\\Python\\data.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet2"]  # workbook.active

# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(r, c).value = "welcome"
#
# workbook.save(file)

# writing multiple date
file = "D:\\programs work space\\Python\\data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet3"]  # workbook.active
sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "smith"

sheet.cell(2, 1).value = 157
sheet.cell(2, 2).value = "John"

sheet.cell(3, 1).value = 223
sheet.cell(3, 2).value = "David"

workbook.save(file)